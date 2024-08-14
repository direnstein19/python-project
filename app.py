from flask import Flask, request, jsonify, send_file, send_file, make_response
import os
import pandas as pd
from fpdf import FPDF
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import requests
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle ,PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import io
import base64
 
app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'})
    if file:
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], file.filename))
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        excel_data = pd.ExcelFile(file_path)
        sheet_columns = {}
        for sheet_name in excel_data.sheet_names:
            sheet = excel_data.parse(sheet_name)  
            column_names = list(sheet.columns)
            sheet_columns[sheet_name] = column_names
        return jsonify({
            'routing': file_path,
            'sheet_count': len(excel_data.sheet_names),
            'sheets_names': excel_data.sheet_names,
            'sheet_columns': sheet_columns
        })
    
@app.route('/build_report', methods=['POST'])
def build_report():
    data = request.json
    file_path = data.get('file_path')
    sheets = data.get('sheets')
    report = {}
    for sheet in sheets:
        sheet_name = sheet.get('name')
        operation = sheet.get('operation')
        columns = sheet.get('columns')
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        result = None
        if operation == 'sum':
            result = df[columns].sum().to_dict()
        elif operation == 'average':
            result = df[columns].mean().to_dict()
        report[sheet_name] = result
    return jsonify(report)

def generate_pdf_report(report_object):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", style="B", size=16)
    pdf.cell(200, 20, "Report Information:", 0, 1, "C")
    pdf.set_font("Arial", size=12)
    pdf.ln(10)
    pdf.set_fill_color(200, 220, 255)
    pdf.set_fill_color(200)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(50, 10, "Sheet", 1, 0, 'C', 1) 
    pdf.cell(140, 10, "Columns and Result", 1, 1, 'C', 1)
    pdf.set_fill_color(220)
    pdf.set_font('Arial', '', 12) 
    for key, value in report_object.items():
        pdf.cell(50, 10, key, 1, 0, 1)
        pdf.cell(140, 10, str(value), 1, 1, 1)  
    pdf_file = "generated_report.pdf"
    pdf.output(pdf_file)  
    return pdf_file  

@app.route('/generate_pdf_report', methods=['POST'])
def generate_pdf_report_endpoint():
    report_data = request.json
    pdf_file = generate_pdf_report(report_data)
    return jsonify({"message": "PDF report generated and downloaded.", "file_name": pdf_file})

@app.route('/get_num_sheets', methods=['POST'])
def get_number_of_sheets_in_excel_data():
    file_path = request.json['file_path']
    return get_number_of_sheets_in_excel(file_path)
def get_number_of_sheets_in_excel(file_path):
    excel_file = pd.ExcelFile(file_path)
    num_sheets = len(excel_file.sheet_names)
    return jsonify({'num_sheets': num_sheets})

@app.route('/calculate_sum', methods=['POST'])
def calculate_sum_of_all_columns_in_all_sheets_data():
    file_path = request.json['file_path']
    return calculate_sum_of_all_columns_in_all_sheets(file_path)
def calculate_sum_of_all_columns_in_all_sheets(file_path):
    excel_file = pd.ExcelFile(file_path)
    sum_columns_data = []
    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(excel_file, sheet_name)
        sum_columns = df.sum().astype(int).apply(lambda x: int(x))
        sum_sheet = int(df.sum().sum())
        sum_columns_data.append({'sheet_name': sheet_name, 'sum_columns': sum_columns.to_dict(), 'sum_sheet': sum_sheet})
    return sum_columns_data

@app.route('/present_column_graph', methods=['POST'])
def present_column_graph_data():
    file_paths = request.json['file_paths']
    return present_column_graph(file_paths)
def present_column_graph(file_paths): 
    files_data = []
    save_images = []   
    for file_path in file_paths:
        sheets_names = []
        sum_sheet = []
        sum_per_sheet = calculate_sum_of_all_columns_in_all_sheets(file_path)
        for sheet in sum_per_sheet:
            sheets_names.append(sheet['sheet_name'])
            sum_sheet.append(sheet['sum_sheet'])
        file_name = os.path.basename(file_path)
        plt.bar(sheets_names, sum_sheet)
        plt.xlabel('Sheets Name')
        plt.ylabel('Sum of all Columns')
        plt.title(f"Sum for {file_name} file")
        plt.savefig(f'sum_graph_{file_name}.png')
        plt.clf()
        files_data.append({'file_name':file_name,
                           'sheets_names':sheets_names,
                           'sum_sheet':sum_sheet,
                           'graph':f'sum_graph_{file_name}.png'})
        save_images.append(f'sum_graph_{file_name}.png')
    return jsonify({'files_data':files_data, 'graphs': f'Column graphs saved as {save_images}.'})

@app.route('/calculate_average_graph', methods=['POST'])
def calculate_average_graph_data():
    file_paths = request.json['file_paths']
    return calculate_average_graph(file_paths)
def calculate_average_graph(file_paths):
    data = {}
    for file_path in file_paths:
        sum_all_sheets = calculate_sum_of_all_columns_in_all_sheets(file_path)
        sum_sheets = 0
        for sheet in sum_all_sheets:
            sum_sheets += sheet['sum_sheet']
        sheets = get_number_of_sheets_in_excel(file_path)
        sheets_file = sheets.json
        sheets_num = sheets_file['num_sheets']
        file_avg = sum_sheets/sheets_num
        file_name = os.path.basename(file_path)
        data[file_name] = file_avg
    plt.figure(figsize=(10, 6))
    plt.bar(data.keys(), data.values())
    plt.xlabel('Files Name')
    plt.ylabel('Average')
    plt.title('Average sums of Excel files')
    plt.savefig('average_graph.png')
    return jsonify({'average':data ,'Average graph saved as':'average_graph.png'})

@app.route('/create_pdf_report_for_data_on_the_requests', methods=['POST'])
def get_data_for_report():
    file_paths = request.json['file_paths']
    data = present_column_graph(file_paths)
    data_json = data.json
    data_files = data_json['files_data']
    average_data = calculate_average_graph(file_paths)
    average = average_data.json
    average_file = average['average']
    for file_data in data_files:
        file_name = file_data["file_name"]
        if file_name in average_file:
            file_data["average_file"] = average_file[file_name]
    return create_pdf_report_for_data(data_files, average['Average graph saved as'])    
def create_pdf_report_for_data(data, average_files_graph):
    pdf_file = "report.pdf"
    doc = SimpleDocTemplate(pdf_file, pagesize=letter)
    story = [] 
    styles = getSampleStyleSheet()
    body_style = styles['BodyText']
    body_style.spaceAfter = 12
    for idx, file_data in enumerate(data):
        if idx > 0:
            story.append(PageBreak()) 
        headings = ['Sheet Name', 'Sum per Sheet']
        sheet_data = [(sheet_name, sum_amt) for sheet_name, sum_amt in zip(file_data['sheets_names'], file_data['sum_sheet'])]
        table = Table([headings] + sheet_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ]))       
        story.append(Paragraph(f"File: {file_data['file_name']}", styles['Title']))
        story.append(table)
        img_path = file_data['graph']
        img_buffer = io.BytesIO()
        img_buffer.write(open(img_path, 'rb').read())
        img_buffer.seek(0)
        space = Spacer(1, 0.5*inch)
        story.append(space)
        styles = getSampleStyleSheet()
        story.append (Paragraph(f"Average File: {file_data['average_file']}", styles['Title']))
        space = Spacer(1, 0.5*inch)
        story.append(space)
        story.append(Paragraph(f"Graph for {file_data['file_name']}", styles['Title']))
        img_data = img_buffer.getvalue()
        img_encoded = base64.b64encode(img_data).decode('utf-8')
        img_src = f"data:image/png;base64,{img_encoded}"
        story.append(Image(img_src, width=300, height=200, hAlign='CENTER'))     
    story.append(PageBreak()) 
    story.append (Paragraph(f"General Graph Of All The Files", styles['Title']))
    image_path = average_files_graph
    average_files = Image(image_path, width=500, height=300, hAlign='CENTER')
    story.append(average_files) 
    doc.build(story) 
    return jsonify({'Report file saved as': pdf_file})

@app.route('/download_pdf_report', methods=['POST'])
def download_report():
    data = get_data_for_report()
    data_file = data.json
    pdf_file_name = data_file['Report file saved as']       
    pdf_data = open(pdf_file_name, 'rb').read()       
    response = make_response(pdf_data)
    response.headers['Content-Disposition'] = f'attachment; filename={pdf_file_name}'
    response.headers['Content-Type'] = 'application/pdf'        
    return response

if __name__ == '__main__':
    app.run(debug=True)